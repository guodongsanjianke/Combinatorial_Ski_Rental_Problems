from collections import defaultdict

import numpy as np

import matplotlib.pyplot as plt
from tqdm import tqdm, trange
import basic_func
from basic_func import opt_line

import openpyxl


wb = openpyxl.Workbook()

import math

n = 6
b = [100, 95, 90, 85, 90, 75]
r = [1, 1.05, 1.10, 1.15, 1.20, 1.25]

STD1 = 1000


Lambda1 = 0.25
num_eval=10000



b_dict = {}
for idx in trange(1, n + 1):
    b_dict[(idx,)] = b[idx - 1]

opt_lines = [opt_line(b=0.0, r=r[0])]
opt_lines.append(opt_line(b=b[n - 1], r=0.0))

opt_result, T = basic_func.get_opt_result1(opt_lines)

# Extend the opt_result
Ext_opt_result = []
s = opt_result[T] / T
for i in range(len(opt_result) - 1):
    if opt_result[i + 1] - opt_result[i] < s:
        Ext_opt_result = opt_result[:i + 1]
        break
while len(Ext_opt_result) <= math.ceil(b[0] / Lambda1):
    if len(Ext_opt_result) > 0:
        Ext_opt_result.append(Ext_opt_result[-1] + s)
    else:
        Ext_opt_result.append(s)

fixed_partitions = [
    [tuple([i])] for i in range(1, n + 1)
]


class CSR_with_ML:
    def __init__(self,Lambda,all_partitions, b_dict, r, Ext_opt_result, T):
        self.results1_for_all_path = None
        self.results2_for_all_path = None

        self.cpt_all_path(Lambda,all_partitions, b_dict, r, Ext_opt_result, T)

    def get_all_path(self, all_partitions, b_dict, r, Ext_opt_result, l):

        u_b, u_r, partitions_list = basic_func.transfer(all_partitions, b_dict, r)
        min_c_max, prob = basic_func.get_min_c_max(all_partitions, u_b, u_r, partitions_list, l, Ext_opt_result)


        t_b, t_r = basic_func.renew_b_r(u_b, u_r, partitions_list, prob)
        p_list, start_day, ALG_x = basic_func.P_list_and_get_ALG(t_b, t_r, min_c_max, l, Ext_opt_result, ret_alg=True)

        p_list_dict = defaultdict(lambda: defaultdict(str))

        extend_T = l + 5
        for i in range(len(p_list)):
            partitions_list[i].p_list = p_list[i]
            p_list_dict[partitions_list[i].p_id][partitions_list[i].split] = basic_func.p_list_extend(p_list[i],
                                                                                                      extend_T,
                                                                                                      start_day[i] - 1)
        all_path = basic_func.All_Path(n=len(all_partitions))
        for index, partition in enumerate(all_partitions):
            # Sort based on the ratio of prices to rent for product combinations
            combined_info = basic_func.combinations_sort(partition, r, b_dict)
            b_, r_, split = map(list, zip(*combined_info))
            p_list = []
            for s in split:
                p_list.append(p_list_dict[index][s])
            all_path.addPath(basic_func.path(prob[index], p_list, b_, r_, split))

        return all_path

    def cpt_all_path(self, Lambda, all_partitions, b_dict, r, Ext_opt_result, T):

        # k = math.floor(Lambda * b[n-1])
        k = int(b[n-1] * Lambda)
        self.results1_for_all_path = self.get_all_path(all_partitions, b_dict, r, Ext_opt_result, k)

        # l = math.ceil(b[0] / Lambda)
        l = int(np.ceil(b[0] / Lambda))
        self.results2_for_all_path = self.get_all_path(all_partitions, b_dict, r, Ext_opt_result, l)
        # print(k,l)

    def get_result(self,case):

        if case==1:
            return self.results1_for_all_path
        elif case==2:
            return self.results2_for_all_path
        return None



def y_sampling(x, std):
    epsilon = np.random.normal(loc=0, scale=std)
    # eta = np.abs(epsilon)
    y = x + epsilon
    y = np.floor(y).astype(int)
    return y

def y_sampling_batch(x_array, std):
    epsilon = np.random.normal(loc=0, scale=std, size=len(x_array))
    y = x_array + epsilon
    return y



def cr_compute( this_partition, opt_result, T, x, y, pre_cpt):
    n = len(this_partition)

    case = 0
    if y >= T:
        case = 1
    else:
        case = 2

    CR = 0.0

    all_path = pre_cpt.get_result(case)
    ALG_x = basic_func.sample_evaluate(all_path, opt_result, end_T=x, ret_cr=False, ret_alg=True)

    if x >= T:
        OPT_x = opt_result[T]
    else:
        OPT_x = opt_result[x]

    CR = ALG_x / OPT_x

    return CR


x_list = []
y_list = []

for STD in range(STD1):
    x_list.append(STD)

for i,Lambda in enumerate([0.25, 0.5, 0.75, 1]):
    y_list.append([])
    pre_cpt = CSR_with_ML(Lambda, fixed_partitions, b_dict, r, Ext_opt_result, T)

    for STD in trange(STD1):

        c_list = []
        for num in range(num_eval):
            x = np.random.randint(1, 301)
            y = y_sampling(x, STD)

            CR = cr_compute(fixed_partitions, opt_result, T, x, y, pre_cpt)


            c_list.append(CR)


        y_list[i].append(np.mean(c_list))



####################################################The second algorithm

def calculate_cr(sample_size=10000, sigma_range=STD1, lambda_values=[0.25]):
    r1, b1 = 1, 100
    r2, b2 = 1.25, 75
    crm_list = []

    for lumda in lambda_values:
        k = int(b2 * lumda)
        l = int(np.ceil(b1 / lumda))


        pdf_A = [((b2 - r2) / b2) ** (k - i) * (r2 / (b2 * (1 - (1 - r2 / b2) ** k))) for i in range(1, k + 1)]


        pdf_B = [((b1 - 1) / b1) ** (l - i) * (1 / (b1 * (1 - (1 - 1 / b1) ** l))) for i in range(1, l + 1)]


        crm = np.zeros(sigma_range)


        for sigma in trange(1, sigma_range + 1):

            total_cr = 0


            for _ in range(sample_size):

                x1 = np.random.randint(1, 301)


                eta1 = np.random.normal(0, sigma)
                y1 = x1 + eta1
                y = int(np.floor(y1))


                A = np.random.choice(np.arange(1, k + 1), p=pdf_A)
                B = np.random.choice(np.arange(1, l + 1), p=pdf_B)


                if y >= b2:
                    if x1 >= b2 and x1 <= A:
                        cr = (x1 * r2) / b2
                    elif x1 >= b2 and x1 > A:
                        cr = ((A - 1) * r2 + b2) / b2
                    elif x1 < b2 and x1 <= A:
                        cr = (x1 * r2) / x1
                    elif x1 < b2 and x1 > A:
                        cr = ((A - 1) * r2 + b2) / x1
                else:
                    if x1 >= b2 and x1 <= B:
                        cr = (x1 * 1) / b2
                    elif x1 >= b2 and x1 > B:
                        cr = ((B - 1) * 1 + b1) / b2
                    elif x1 < b2 and x1 <= B:
                        cr = x1 / x1  # 等于1
                    elif x1 < b2 and x1 > B:
                        cr = ((B - 1) * 1 + b1) / x1


                total_cr += cr


            crm[sigma - 1] = total_cr / sample_size

        crm_list.append(crm)


        plt.plot(range(1, sigma_range + 1), crm, label=f'λ = {lumda}')

    plt.title('Competition ratio using PDF random sampling')
    plt.xlabel('σ')
    plt.ylabel('competition ratio')
    plt.legend()
    plt.show()

    return crm_list

crm_list = calculate_cr(sample_size=10000, sigma_range=STD1, lambda_values=[0.25, 0.5, 0.75, 1])




x_list_adj = list(range(1, STD1+1))  # 1 to 300

assert len(x_list) == STD1, "x_list length should be 300"
assert len(y_list) == 4, "y_list should have 4 Lambda values"


# crm_list = [...]  # List of CR arrays for Lambda = [0.25, 0.5, 0.75, 1]
assert len(crm_list) == 4, "crm_list should have 4 Lambda values"
assert len(crm_list[0]) == STD1, "crm_list entries should have length 300"


lambda_labels = [0.25, 0.5, 0.75, 1.0]


plt.figure(figsize=(10, 6))


for i, cr_values in enumerate(y_list):
    plt.plot(x_list_adj[:len(cr_values)], cr_values, label=f'CSR_with_ML, λ={lambda_labels[i]}', linestyle='-', linewidth=2)


for i, crm in enumerate(crm_list):
    plt.plot(x_list_adj, crm, label=f'calculate_cr, λ={lambda_labels[i]}', linestyle='--', linewidth=2)


plt.title('Comparison of Competitive Ratios (CSR_with_ML vs calculate_cr)', fontsize=14)
plt.xlabel('STD (σ)', fontsize=12)
plt.ylabel('Competitive Ratio (CR)', fontsize=12)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=10)
plt.tight_layout()


plt.savefig('comparison_cr.png', dpi=300)
plt.show()

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Remove the default sheet created by openpyxl
wb.remove(wb.active)

# Iterate over each Lambda value
for idx, lambda_val in enumerate(lambda_labels):
    # Create a new sheet for this Lambda value
    sheet = wb.create_sheet(title=f'Lambda_{lambda_val}')

    # Write headers
    sheet['A1'] = 'STD'
    sheet['B1'] = 'CSR_with_ML_CR'
    sheet['C1'] = 'calculate_cr_CR'

    # Write data
    for row, std in enumerate(x_list, start=2):  # Start from row 2 (after header)
        sheet[f'A{row}'] = std
        sheet[f'B{row}'] = y_list[idx][row - 2]  # y_list[idx] contains CR for this Lambda
        sheet[f'C{row}'] = crm_list[idx][row - 2]  # crm_list[idx] contains CR for this Lambda

# Save the workbook
excel_file = 'competitive_ratios.xlsx'
wb.save(excel_file)
print(f"Data saved to {excel_file}")