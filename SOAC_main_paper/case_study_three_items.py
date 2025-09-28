from CSR import csr,evaluate
from upgrading import csr_upgrading
import openpyxl
import argparse

parser = argparse.ArgumentParser(description="")
parser.add_argument('--eval_T', type=int, default=None, help='Maximum number of test days')
# eval_T = int(np.max(2 * np.array(b) / np.array(r)))  copy from CSR.py line. 86
parser.add_argument('--num_eval', type=int, default=100000, help='Number of samples per day')
args = parser.parse_args()

wb = openpyxl.Workbook()

# Order
# Word2021 / Excel2021 / PowerPoint 2021

b = [149.99, 149.99, 149.99]
r = [0.30, 0.80, 0.50]

b_12 = 229.99
b_13 = 229.99
b_23 = 229.99
b_123 = 329.99

# Define the price of a combination of items
b_dict = {
    (): 0,
    (1,): 149.99,
    (2,): 149.99,
    (3,): 149.99,
    (1, 2): b_12,
    (1, 3): b_13,
    (2, 3): b_23,
    (1, 2, 3): b_123,
}

#################################
# upgrading and non-upgrading   #
#################################

result = csr(b, r, b_dict=b_dict,eval_T=args.eval_T)

CR = result['CR']
print(f'Non-upgrading, CR: {CR}')

CR_upgrading = csr_upgrading(b, r, b_dict=b_dict)
print(f'Upgrading, CR: {CR_upgrading}')


##################################################################
print('########################')
print('Non-upgrading, the probability of 2 layers')

all_path,opt_result,eval_T = result['all_path'],result['opt_result'],result['eval_T']

sheet1 = wb.active
sheet1.title = "Non-upgrading"
sheet1.append(["Partitions ofItems", "The probability of the first layer", "The probability of the second layer","CR"])

for path in all_path.paths:

    print(str(path.items_split) + ":")
    print(f'The probability of the first layer:{path.prob}')
    # print(f'The probability of the second layer:')
    # print(path.p_list)
    sheet1.append([str(path.items_split), path.prob, str(path.p_list)])

eval_cr = evaluate(all_path, opt_result, eval_T,num_evaluate=args.num_eval)
sheet1.append(["CR",f"{CR}","evaluate_cr",str(eval_cr)])
##################################################################


####################################
# Buy together and buy separately  #
####################################


##################################################################
print('########################')
print('Buy together')

result1 = csr(b, r, fixed_partitions=[[(1, 2, 3)]], b_dict=b_dict,eval_T=args.eval_T)
CR1 = result1['CR']
all_path1,opt_result1,eval_T1 = result1['all_path'],result1['opt_result'],result1['eval_T']

print(f'CR: {CR1}')

sheet2 = wb.create_sheet(title="Buy together")
sheet2.append(["Partitions of Items", "The probability of the first layer", "The probability of the second layer"])

for path in all_path1.paths:
    print(str(path.items_split) + ":")
    print(f'The probability of the first layer: {path.prob}')
    # print(f'The probability of the second layer:')
    # print(path.p_list)
    sheet2.append([str(path.items_split), path.prob, str(path.p_list)])

eval_cr1 = evaluate(all_path1, opt_result, eval_T,num_evaluate=args.num_eval)
sheet2.append(["CR",f"{CR1}","evaluate_cr",str(eval_cr1)])
##################################################################


##################################################################
print('########################')
print('Buy separately')

result2 = csr(b, r, fixed_partitions=[[(1,), (2,), (3,)]], b_dict=b_dict,eval_T=args.eval_T)
CR2 = result2['CR']
all_path2,opt_result2,eval_T2 = result2['all_path'],result2['opt_result'],result2['eval_T']

print(f'CR: {CR2}')

sheet3 = wb.create_sheet(title="Buy separately")
sheet3.append(["Partitions of Items", "The probability of the first layer", "The probability of the second layer"])

for path in all_path2.paths:
    print(str(path.items_split) + ":")
    print(f'The probability of the first layer: {path.prob}')
    # print(f'The probability of the second layer:')
    # print(path.p_list)
    sheet3.append([str(path.items_split), path.prob, str(path.p_list)])

eval_cr2 = evaluate(all_path2, opt_result, eval_T,num_evaluate=args.num_eval)
sheet3.append(["CR",f"{CR2}","evaluate_cr",str(eval_cr2)])
##################################################################



##################################################################
print('########################')
print('path 2')

result3 = csr(b, r, fixed_partitions=[[(1, 2), (3,)]], b_dict=b_dict)
CR3 = result3['CR']
all_path3,opt_result3,eval_T3 = result3['all_path'],result3['opt_result'],result3['eval_T']

print(f'CR: {CR3}')

sheet4 = wb.create_sheet(title="12_3")
sheet4.append(["Partitions of Items", "The probability of the first layer", "The probability of the second layer"])

for path in all_path3.paths:
    print(str(path.items_split) + ":")
    print(f'The probability of the first layer: {path.prob}')

eval_cr3 = evaluate(all_path3, opt_result, eval_T, num_evaluate=args.num_eval)
sheet4.append(["CR",f"{CR3}","evaluate_cr",str(eval_cr3)])

##################################################################


##################################################################
print('########################')
print('path 3')

result4 = csr(b, r, fixed_partitions=[[(2,), (1, 3)]], b_dict=b_dict)
CR4 = result4['CR']
all_path4,opt_result4,eval_T4 = result4['all_path'],result4['opt_result'],result4['eval_T']

print(f'CR: {CR4}')

sheet5 = wb.create_sheet(title="2_13")
sheet5.append(["Partitions of Items", "The probability of the first layer", "The probability of the second layer"])

for path in all_path4.paths:
    print(str(path.items_split) + ":")
    print(f'The probability of the first layer: {path.prob}')

eval_cr4 = evaluate(all_path4, opt_result, eval_T,num_evaluate=args.num_eval)
sheet5.append(["CR",f"{CR4}","evaluate_cr",str(eval_cr4)])
##################################################################


##################################################################
print('########################')
print('path 4')

result5 = csr(b, r, fixed_partitions=[[(2, 3), (1,)]], b_dict=b_dict)
CR5 = result5['CR']
all_path5,opt_result5,eval_T5 = result5['all_path'],result5['opt_result'],result5['eval_T']

print(f'CR: {CR5}')

sheet6 = wb.create_sheet(title="23_1")
sheet6.append(["Partitions of Items", "The probability of the first layer", "The probability of the second layer"])

for path in all_path5.paths:
    print(str(path.items_split) + ":")
    print(f'The probability of the first layer: {path.prob}')

eval_cr5 = evaluate(all_path5, opt_result, eval_T,num_evaluate=args.num_eval)
sheet6.append(["CR",f"{CR5}","evaluate_cr",str(eval_cr5)])
##################################################################


wb.save("path_data.xlsx")
print("Data has been successfully written to path_data.xlsx")
